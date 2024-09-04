import requests # type: ignore
import sys
import openpyxl # type: ignore
from openpyxl.styles import PatternFill, Alignment # type: ignore

def extrair_estatisticas_jogador(jogador, total_time):
    role = jogador.get("role", "NONE")
    if role == "NONE":
        role = "Jungle"
        
    estatisticas = {
        "riot_id_nome": jogador.get("riotIdGameName", "Desconhecido"),
        "campeão": jogador["championName"],
        "abates": jogador["kills"],
        "mortes": jogador["deaths"],
        "assistências": jogador["assists"],
        "farm": jogador["totalMinionsKilled"],
        "ouro": jogador["goldEarned"],
        "dano": jogador["totalDamageDealtToChampions"],
        "sentinelas": jogador["wardsPlaced"],
        "dano_sofrido": jogador["totalDamageTaken"],
        "controle_de_grupos": jogador["totalTimeCCDealt"],
        "role": role  # Adicionando role com substituição
    }
    # Calculando as novas estatísticas
    estatisticas["dano_por_minuto"] = estatisticas["dano"] / (total_time / 60)
    estatisticas["gold_por_minuto"] = estatisticas["ouro"] / (total_time / 60)
    estatisticas["ward_por_minuto"] = estatisticas["sentinelas"] / (total_time / 60)
    
    # Calcular KDA
    deaths = jogador["deaths"]
    estatisticas["kda"] = (estatisticas["abates"] + estatisticas["assistências"]) / deaths if deaths > 0 else (estatisticas["abates"] + estatisticas["assistências"])

    return estatisticas

def calcular_percentuais(time_100, time_200):
    total_abates = sum(jogador['abates'] for jogador in time_100 + time_200)
    total_ouro = sum(jogador['ouro'] for jogador in time_100 + time_200)
    total_dano = sum(jogador['dano'] for jogador in time_100 + time_200)
    
    for jogador in time_100 + time_200:
        jogador["percentual_kills"] = (jogador['abates'] / total_abates) * 100 if total_abates > 0 else 0
        jogador["percentual_gold"] = (jogador['ouro'] / total_ouro) * 100 if total_ouro > 0 else 0
        jogador["percentual_dano"] = (jogador['dano'] / total_dano) * 100 if total_dano > 0 else 0

def separar_times(dados):
    time_100 = []
    time_200 = []
    
    total_time = dados["info"]["gameDuration"]
    
    for jogador in dados["info"]["participants"]:
        estatisticas_jogador = extrair_estatisticas_jogador(jogador, total_time)
        if jogador["teamId"] == 100:
            time_100.append(estatisticas_jogador)
        else:
            time_200.append(estatisticas_jogador)
    
    calcular_percentuais(time_100, time_200)
    
    return time_100, time_200

def adicionar_estatisticas_ao_excel(ws, estatisticas, row_start):
    for i, jogador in enumerate(estatisticas):
        ws[f'A{row_start + i}'] = jogador['role']  # Adicionar role na coluna A
        ws[f'B{row_start + i}'] = jogador['riot_id_nome']
        ws[f'C{row_start + i}'] = jogador['campeão']
        ws[f'D{row_start + i}'] = jogador['abates']
        ws[f'E{row_start + i}'] = jogador['mortes']
        ws[f'F{row_start + i}'] = jogador['assistências']
        ws[f'G{row_start + i}'] = jogador['kda']
        ws[f'H{row_start + i}'] = jogador['farm']
        ws[f'I{row_start + i}'] = jogador['ouro']
        ws[f'J{row_start + i}'] = jogador['dano']
        ws[f'K{row_start + i}'] = jogador['sentinelas']
        ws[f'L{row_start + i}'] = jogador['dano_sofrido']
        ws[f'M{row_start + i}'] = jogador['controle_de_grupos']
        ws[f'N{row_start + i}'] = jogador['percentual_kills']
        ws[f'O{row_start + i}'] = jogador['percentual_gold']
        ws[f'P{row_start + i}'] = jogador['percentual_dano']
        ws[f'Q{row_start + i}'] = jogador['dano_por_minuto']
        ws[f'R{row_start + i}'] = jogador['gold_por_minuto']
        ws[f'S{row_start + i}'] = jogador['ward_por_minuto']

def processar_partida(api_key, match_id, ws, row_start):
    region = 'americas'
    formatted_match_id = f'BR1_{match_id}'
    params = {'api_key': api_key}
    base_url = f'https://{region}.api.riotgames.com/lol/match/v5/matches/{formatted_match_id}'

    response = requests.get(base_url, params=params)

    if response.status_code == 200:
        match_data = response.json()
        time_100, time_200 = separar_times(match_data)

        # Adicionando estatísticas ao Excel
        adicionar_estatisticas_ao_excel(ws, time_100, row_start)
        adicionar_estatisticas_ao_excel(ws, time_200, row_start + len(time_100))

        return len(time_100) + len(time_200)
    else:
        print(f"Erro ao obter informações da partida {formatted_match_id}: {response.status_code}")
        print(response.text)
        return 0

def main():
    if len(sys.argv) < 3:
        print("Uso: python script.py <API_KEY> <arquivo_excel> <MATCH_ID_1> <MATCH_ID_2> ... <MATCH_ID_N>")
        sys.exit(1)

    api_key = sys.argv[1]
    arquivo_excel = sys.argv[2]
    match_ids = sys.argv[3:]

    # Criar um novo arquivo Excel e configurar a planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estatísticas das Partidas"

    # Adicionar a célula mesclada para "Análise de Performance"
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws.merge_cells('A1:U1')
    ws['A1'] = "Análise de Performance"
    ws['A1'].fill = fill_red
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Adicionar cabeçalhos das colunas diretamente abaixo
    headers = [
        "Role", "Riot ID Nome", "Campeão", "Abates", "Mortes", "Assistências", "KDA", "Farm", "Ouro", 
        "Dano", "Sentinelas", "Dano Sofrido", "Controle de Grupos", "% Kills", "% Ouro", "% Dano", 
        "Dano por Minuto", "Gold por Minuto", "Ward por Minuto"
    ]
    ws.append(headers)
    
    # Adicionar a célula para o cabeçalho das roles
    ws['A2'] = "Role"
    ws['A2'].fill = fill_red
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar a largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20

    row_start = 3  # Começando na linha 3 (linha 2 é o cabeçalho das roles e linha 1 é a análise de performance)

    for match_id in match_ids:
        num_rows = processar_partida(api_key, match_id, ws, row_start)
        row_start += num_rows

    # Salvar o novo arquivo Excel
    wb.save(arquivo_excel)
    print(f"Estatísticas adicionadas ao arquivo {arquivo_excel} com sucesso.")

if __name__ == "__main__":
    main()
